"""Generate sample.xlsx fixture for tests."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path


def create_sample():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # Header row (bold)
    headers = ["Method", "Accuracy", "F1 Score", "Params (M)"]
    bold_font = Font(bold=True)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    data = [
        ["Baseline", 0.823, 0.801, 12.5],
        ["Ours", 0.891, 0.876, 8.3],
        ["Ours+Aug", 0.912, 0.895, 8.3],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal="center")

    # Bold best result
    ws.cell(row=4, column=2).font = Font(bold=True)
    ws.cell(row=4, column=3).font = Font(bold=True)

    # Average row with all values
    ws.cell(row=5, column=1, value="Average")
    ws.cell(row=5, column=1).font = Font(italic=True)
    ws.cell(row=5, column=2, value=0.875)
    ws.cell(row=5, column=3, value=0.857)
    ws.cell(row=5, column=4, value=9.7)

    out = Path(__file__).parent / "fixtures" / "sample.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Created {out}")


if __name__ == "__main__":
    create_sample()
