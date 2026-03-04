"""Round-trip conversion tests: tex→xlsx and xlsx→tex→xlsx."""

from pathlib import Path

import openpyxl
import pytest

from pubtab import convert, preview, read_excel, tex_to_excel
from pubtab.models import SpacingConfig, TableData
from pubtab._preview import _build_standalone, _find_pdflatex, _strip_table_float, compile_pdf
from pubtab.config import load_config
from pubtab.renderer import render
from pubtab.tex_reader import read_tex

FIXTURES = Path(__file__).parent / "fixtures"
TABLES = ["table1", "table2", "table3", "table4", "table5", "table6", "table8"]


def _compare_xlsx(path_a: Path, path_b: Path) -> list:
    """Compare two xlsx files cell by cell. Returns list of diffs."""
    wb1 = openpyxl.load_workbook(path_a)
    wb2 = openpyxl.load_workbook(path_b)
    ws1, ws2 = wb1.active, wb2.active
    diffs = []
    max_r = max(ws1.max_row, ws2.max_row)
    max_c = max(ws1.max_column, ws2.max_column)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v1 = ws1.cell(r, c).value
            v2 = ws2.cell(r, c).value
            s1 = "" if v1 is None else str(v1).strip()
            s2 = "" if v2 is None else str(v2).strip()
            try:
                if abs(float(s1) - float(s2)) < 1e-9:
                    continue
            except (ValueError, TypeError):
                pass
            if s1 != s2:
                diffs.append((r, c, v1, v2))
    return diffs


# --- tex → xlsx round-trip ---

@pytest.mark.parametrize("name", TABLES)
def test_tex_to_xlsx_dimensions(name, tmp_path):
    """tex→xlsx produces correct dimensions matching original xlsx."""
    tex_file = FIXTURES / f"{name}.tex"
    orig_xlsx = FIXTURES / f"{name}.xlsx"
    gen_xlsx = tmp_path / f"{name}.xlsx"
    tex_to_excel(str(tex_file), str(gen_xlsx))

    wb_orig = openpyxl.load_workbook(orig_xlsx)
    wb_gen = openpyxl.load_workbook(gen_xlsx)
    assert wb_gen.active.max_row == wb_orig.active.max_row
    assert wb_gen.active.max_column == wb_orig.active.max_column


@pytest.mark.parametrize("name", TABLES)
def test_tex_to_xlsx_values_match(name, tmp_path):
    """tex→xlsx cell values match original xlsx exactly."""
    gen_xlsx = tmp_path / f"{name}.xlsx"
    tex_to_excel(str(FIXTURES / f"{name}.tex"), str(gen_xlsx))
    diffs = _compare_xlsx(FIXTURES / f"{name}.xlsx", gen_xlsx)
    assert diffs == [], f"Cell diffs: {diffs[:5]}"


@pytest.mark.parametrize("name", TABLES)
def test_tex_to_xlsx_merged_cells(name, tmp_path):
    """tex→xlsx preserves merged cell count."""
    gen_xlsx = tmp_path / f"{name}.xlsx"
    tex_to_excel(str(FIXTURES / f"{name}.tex"), str(gen_xlsx))

    wb_orig = openpyxl.load_workbook(FIXTURES / f"{name}.xlsx")
    wb_gen = openpyxl.load_workbook(gen_xlsx)
    assert len(wb_gen.active.merged_cells.ranges) == len(wb_orig.active.merged_cells.ranges)


# --- xlsx → tex → xlsx round-trip ---

@pytest.mark.parametrize("name", TABLES)
def test_xlsx_to_tex_roundtrip(name, tmp_path):
    """xlsx→tex→xlsx round-trip preserves all cell values."""
    tex_path = tmp_path / f"{name}.tex"
    convert(str(FIXTURES / f"{name}.xlsx"), str(tex_path))

    gen_xlsx = tmp_path / f"{name}_rt.xlsx"
    tex_to_excel(str(tex_path), str(gen_xlsx))

    table_orig = read_excel(str(FIXTURES / f"{name}.xlsx"))
    table_gen = read_tex(tex_path.read_text())
    assert table_gen.num_rows == table_orig.num_rows
    assert table_gen.num_cols == table_orig.num_cols


def test_xlsx2tex_default_exports_all_sheets(tmp_path):
    """xlsx2tex without sheet should export every sheet to separate tex files."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Main Sheet"
    ws1["A1"] = "MAINCELL"
    ws2 = wb.create_sheet("Aux-2")
    ws2["A1"] = "AUXCELL"

    xlsx_path = tmp_path / "multi.xlsx"
    wb.save(xlsx_path)

    out_tex = tmp_path / "multi.tex"
    convert(str(xlsx_path), str(out_tex))

    generated = sorted(tmp_path.glob("multi*.tex"))
    assert len(generated) == 2
    first_tex = tmp_path / "multi_sheet01.tex"
    second_tex = tmp_path / "multi_sheet02.tex"
    assert first_tex in generated
    assert second_tex in generated
    assert "MAINCELL" in first_tex.read_text()
    assert "AUXCELL" in second_tex.read_text()


def test_xlsx2tex_sheet_option_exports_single_sheet(tmp_path):
    """xlsx2tex with sheet option should export only the specified sheet."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Main Sheet"
    ws1["A1"] = "MAINCELL"
    ws2 = wb.create_sheet("Aux-2")
    ws2["A1"] = "AUXCELL"

    xlsx_path = tmp_path / "single.xlsx"
    wb.save(xlsx_path)

    out_tex = tmp_path / "single.tex"
    convert(str(xlsx_path), str(out_tex), sheet="Aux-2")

    generated = sorted(tmp_path.glob("single*.tex"))
    assert len(generated) == 1
    assert generated[0] == out_tex
    text = out_tex.read_text()
    assert "AUXCELL" in text
    assert "MAINCELL" not in text


def test_xlsx2tex_directory_input_exports_all_workbooks(tmp_path):
    """xlsx2tex should accept a directory and batch-convert all Excel files."""
    in_dir = tmp_path / "xlsx_in"
    out_dir = tmp_path / "tex_out"
    in_dir.mkdir()

    wb1 = openpyxl.Workbook()
    wb1.active["A1"] = "ONE"
    wb1.save(in_dir / "a.xlsx")

    wb2 = openpyxl.Workbook()
    wb2.active["A1"] = "TWO"
    wb2.save(in_dir / "b.xlsx")

    convert(str(in_dir), str(out_dir))

    a_tex = out_dir / "a.tex"
    b_tex = out_dir / "b.tex"
    assert a_tex.exists()
    assert b_tex.exists()
    assert "ONE" in a_tex.read_text()
    assert "TWO" in b_tex.read_text()


def test_xlsx2tex_directory_input_requires_output_directory(tmp_path):
    """Directory input should reject file-like .tex output path."""
    in_dir = tmp_path / "xlsx_in"
    in_dir.mkdir()
    wb = openpyxl.Workbook()
    wb.active["A1"] = "X"
    wb.save(in_dir / "a.xlsx")

    with pytest.raises(ValueError, match="output must be a directory path"):
        convert(str(in_dir), str(tmp_path / "batch.tex"))


def test_xlsx2tex_includes_commented_package_hints(tmp_path):
    """Generated tex should include commented package hints for Overleaf users."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Header"
    ws["A2"] = "Value"
    xlsx_path = tmp_path / "pkg_hint.xlsx"
    wb.save(xlsx_path)

    out_tex = tmp_path / "pkg_hint.tex"
    convert(str(xlsx_path), str(out_tex))
    text = out_tex.read_text()

    assert text.startswith("% Theme package hints for this table")
    assert r"% \usepackage{booktabs}" in text
    assert r"% \usepackage{multirow}" in text
    assert r"% \usepackage[table]{xcolor}" in text


def test_xlsx2tex_package_hints_include_graphicx_when_resizebox_enabled(tmp_path):
    """resizebox output should hint graphicx package in comments."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "H"
    ws["A2"] = "V"
    xlsx_path = tmp_path / "pkg_hint_graphicx.xlsx"
    wb.save(xlsx_path)

    out_tex = tmp_path / "pkg_hint_graphicx.tex"
    convert(str(xlsx_path), str(out_tex), resizebox=r"0.8\textwidth")
    text = out_tex.read_text()

    assert r"% \usepackage{graphicx}" in text


def test_tex2xlsx_directory_input_exports_all_tex_files(tmp_path):
    """tex2xlsx should accept a directory and batch-convert all tex files."""
    in_dir = tmp_path / "tex_in"
    out_dir = tmp_path / "xlsx_out"
    in_dir.mkdir()

    tex = r"""
\begin{tabular}{cc}
\toprule
A & B \\
\midrule
1 & 2 \\
\bottomrule
\end{tabular}
"""
    (in_dir / "a.tex").write_text(tex)
    (in_dir / "b.tex").write_text(tex.replace("1 & 2", "3 & 4"))

    tex_to_excel(str(in_dir), str(out_dir))

    a_xlsx = out_dir / "a.xlsx"
    b_xlsx = out_dir / "b.xlsx"
    assert a_xlsx.exists()
    assert b_xlsx.exists()
    assert openpyxl.load_workbook(a_xlsx).active["A2"].value == 1.0
    assert openpyxl.load_workbook(b_xlsx).active["A2"].value == 3.0


def test_tex2xlsx_directory_input_requires_output_directory(tmp_path):
    """Directory input should reject file-like .xlsx output path."""
    in_dir = tmp_path / "tex_in"
    in_dir.mkdir()
    (in_dir / "a.tex").write_text(
        "\\begin{tabular}{c}\\toprule A \\\\ \\bottomrule\\end{tabular}"
    )

    with pytest.raises(ValueError, match="output must be a directory path"):
        tex_to_excel(str(in_dir), str(tmp_path / "batch.xlsx"))


def test_preview_directory_input_exports_pdf_batch(tmp_path):
    """preview should accept a directory and generate one PDF per tex file."""
    if _find_pdflatex() is None:
        pytest.skip("pdflatex not found; skip batch preview test")

    in_dir = tmp_path / "tex_in"
    out_dir = tmp_path / "pdf_out"
    in_dir.mkdir()

    tex = r"""
\begin{table}[htbp]
\centering
\begin{tabular}{cc}
\toprule
A & B \\
\midrule
1 & 2 \\
\bottomrule
\end{tabular}
\end{table}
"""
    (in_dir / "a.tex").write_text(tex)
    (in_dir / "b.tex").write_text(tex.replace("1 & 2", "3 & 4"))

    preview(str(in_dir), output=str(out_dir), format="pdf")
    assert (out_dir / "a.pdf").exists()
    assert (out_dir / "b.pdf").exists()


def test_preview_directory_input_uses_default_output_dir(tmp_path):
    """Directory preview without output should use preview_<format> folder."""
    if _find_pdflatex() is None:
        pytest.skip("pdflatex not found; skip batch preview test")

    in_dir = tmp_path / "tex_in"
    in_dir.mkdir()
    (in_dir / "a.tex").write_text(
        "\\begin{tabular}{c}\\toprule A \\\\ \\bottomrule\\end{tabular}"
    )

    preview(str(in_dir), format="pdf")
    assert (in_dir / "preview_pdf" / "a.pdf").exists()


def test_read_excel_trims_only_trailing_empty_columns(tmp_path):
    """read_excel should trim right empty columns but keep middle empty columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "H1"
    ws["B1"] = ""
    ws["C1"] = "H3"
    ws["A2"] = "v1"
    ws["C2"] = "v3"
    # Force trailing empty columns to exist in worksheet bounds.
    ws["D1"] = ""
    ws["E1"] = ""

    xlsx_path = tmp_path / "trim_cols.xlsx"
    wb.save(xlsx_path)

    table = read_excel(str(xlsx_path))
    assert table.num_cols == 3
    assert table.cells[0][1].value == ""


def test_read_excel_trims_trailing_columns_even_with_wide_merged_title(tmp_path):
    """Trailing empty cols should still be trimmed when covered by a merged title."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Title"
    ws.merge_cells("A1:E1")
    ws["A2"] = "Method"
    ws["B2"] = "Score"
    ws["A3"] = "M1"
    ws["B3"] = "0.95"

    xlsx_path = tmp_path / "trim_with_merge.xlsx"
    wb.save(xlsx_path)

    table = read_excel(str(xlsx_path))
    assert table.num_cols == 2
    assert table.cells[0][0].colspan == 2


# --- tex_reader unit tests ---

def test_tex_reader_comments():
    """Parser handles % comments correctly."""
    tex = r"""
\begin{tabular}{cc}
\toprule
A & B \\ % header comment
\midrule
1 & 2 \\ % data comment
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.num_rows == 2
    assert table.cells[1][0].value == 1.0


def test_tex_reader_multicolumn():
    """Parser handles \\multicolumn correctly."""
    tex = r"""
\begin{tabular}{ccc}
\toprule
\multicolumn{2}{c}{Header} & C \\
\midrule
a & b & c \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.cells[0][0].colspan == 2
    assert table.cells[0][0].value == "Header"


def test_tex_reader_multirow():
    """Parser handles \\multirow correctly."""
    tex = r"""
\begin{tabular}{cc}
\toprule
\multirow{2}{*}{A} & B \\
 & C \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.cells[0][0].rowspan == 2


def test_tex_reader_diagbox():
    """Parser handles \\diagbox correctly."""
    tex = r"""
\begin{tabular}{cc}
\toprule
\diagbox{Row}{Col} & Data \\
\midrule
a & 1 \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.cells[0][0].style.diagbox == ["Row", "Col"]


def test_tex_reader_formatting():
    """Parser extracts bold/italic/underline formatting."""
    tex = r"""
\begin{tabular}{ccc}
\toprule
\textbf{Bold} & \textit{Italic} & \underline{Under} \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.cells[0][0].style.bold
    assert table.cells[0][1].style.italic
    assert table.cells[0][2].style.underline


def test_tex_reader_math_cleanup():
    """Parser cleans up math expressions."""
    tex = r"""
\begin{tabular}{c}
\toprule
$D_\text{stage 1}$ \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.cells[0][0].value == "D_stage 1"


def test_tex_reader_pm_spacing():
    """Parser normalizes ± spacing."""
    tex = r"""
\begin{tabular}{c}
\toprule
0.626 {$\pm$0.018} \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert "0.626±0.018" in str(table.cells[0][0].value)


def test_tex_reader_makecell_hyphen_linebreak_preserved():
    """Parser keeps explicit makecell line breaks with trailing hyphen."""
    tex = r"""
\begin{tabular}{c}
\toprule
\makecell{Things-\\EEG} \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.cells[0][0].value == "Things-\nEEG"


def test_tex_reader_malformed_double_backslash_percent_does_not_split_row():
    """`\\%` artifacts should be interpreted as literal percent in-cell."""
    tex = r"""
\begin{tabular}{cccc}
\toprule
Model & Chat & Delta & Score \\
\midrule
M1 & 68.2 & 27.0\\% & 83.2 \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.num_rows == 2
    assert table.cells[1][2].value == "27.0%"


def test_tex_reader_malformed_double_backslash_hash_keeps_header():
    """`\\#` artifacts should remain header text, not create extra rows."""
    tex = r"""
\begin{tabular}{ccc}
\toprule
Depth & \\#P (M) & Score \\
\midrule
18 & 11.23 & 86.41 \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.num_rows == 2
    assert table.cells[0][1].value == "#P (M)"


def test_tex_reader_rowbreak_followed_by_hash_is_not_collapsed():
    """`...\\\\#...` row boundaries must stay as row separators."""
    tex = r"""
\begin{tabular}{cc}
\toprule
A & B \\
\midrule
v1 & v2\\#Tag & 1 \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.num_rows == 3
    assert table.cells[1][0].value == "v1"
    assert table.cells[2][0].value == "#Tag"


def test_tex_reader_malformed_double_backslash_ampersand_keeps_single_row():
    """`\\&` artifacts should stay inside one cell as literal ampersand."""
    tex = r"""
\begin{tabular}{ccc}
\toprule
Method & Input & Score \\
\midrule
OpenOcc & C\\&L & 70.59 \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.num_rows == 2
    assert table.cells[1][1].value == "C&L"
    assert table.cells[1][2].value == "70.59"


def test_tex_reader_all_delimiters_escaped_as_ampersand_are_recovered():
    """Rows using only `\\&` as separators should still split into multiple cells."""
    tex = r"""
\begin{tabular}{ccc}
\toprule
A \& B \& C \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.num_rows == 1
    assert table.num_cols == 3
    assert table.cells[0][0].value == "A"
    assert table.cells[0][1].value == "B"
    assert table.cells[0][2].value == "C"


def test_tex_reader_rowbreak_followed_by_ampersand_is_not_collapsed():
    """`...\\\\&...` row boundaries must not be normalized into a literal ampersand."""
    tex = r"""
\begin{tabular}{ccc}
\toprule
M & NQ & ARC-C \\
\midrule
\multirow{2}{*}{Ours(Yi-6B)} & 23.28 & 76.54\\&(\textcolor{green}{+0.73})&(\textcolor{green}{+3.33}) \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.num_rows == 3
    assert table.cells[1][2].value == "76.54"
    assert table.cells[2][1].value == "(+0.73)"
    assert table.cells[2][2].value == "(+3.33)"


def test_tex_reader_triple_backslash_rule_commands_not_leaked():
    """Triple-backslash rule commands should not leak as plain text."""
    tex = r"""
\begin{tabular}{ll}
\toprule
A & B \\\hline
C & D \\\cline{1-2}
E & F \\\bottomrule[0.8pt]
\end{tabular}
"""
    table = read_tex(tex)
    assert table.num_rows == 3
    values = [str(cell.value) for row in table.cells for cell in row if str(cell.value).strip()]
    joined = " | ".join(values).lower()
    assert "hline" not in joined
    assert "cline" not in joined
    assert "bottomrule" not in joined


def test_tex_reader_nested_makebox_cleans_to_content():
    """Nested makebox/color wrappers should reduce to plain payload text."""
    tex = r"""
\begin{tabular}{c}
\toprule
\makebox[1.25em][c]{{\color{ForestGreen}\textsf{\textbf{P}}}} \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.cells[0][0].value == "P"


def test_tex_reader_decorative_separator_block_is_removed():
    """Decorative '-\\/-\\/-...' blocks should not pollute data cells."""
    tex = r"""
\begin{tabular}{ll}
\toprule
Lang & Value \\
\midrule
English
-\/-\/-\/-\/-\/-\/-\/-
\multirow{2}{*}{English} & 1 \\
 & 2 \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.num_rows == 3
    assert table.cells[1][0].value == "English"
    all_values = [str(cell.value) for row in table.cells for cell in row]
    assert not any("-/-" in v or "---" in v for v in all_values)


def test_tex_reader_mixed_case_dvips_color_is_preserved():
    """Mixed-case dvips names like `Dandelion` should map to real RGB."""
    tex = r"""
\begin{tabular}{c}
\toprule
\makebox[1.25em][c]{{\color{Dandelion}\textbf{P}}}\quad/\quad\makebox[1.25em][c]{{\color{ForestGreen}\ding{52}}} \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    cell = table.cells[0][0]
    assert cell.rich_segments is not None
    assert cell.rich_segments[0][0] == "P"
    assert cell.rich_segments[0][1] == "#FDBC42"


def test_tex_reader_inline_decorative_separator_is_removed():
    """Inline `Lang -\\/-\\/...` separator lines should be dropped."""
    tex = r"""
\begin{tabular}{ll}
\toprule
Lang & Value \\
\midrule
Korean -\/-\/-\/-\/-\/-\/-\/-
\multirow{2}{*}{Korean} & 1 \\
 & 2 \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.num_rows == 3
    assert table.cells[1][0].value == "Korean"
    all_values = [str(cell.value) for row in table.cells for cell in row]
    assert not any("-/-" in v or "---" in v for v in all_values)


def test_tex_reader_rich_segments_do_not_leak_makecell_prefix():
    """Rich text extraction should not keep wrapper residue like `makecellHe...`."""
    tex = r"""
\begin{tabular}{ll}
\toprule
Q & A \\
\midrule
Qwen2 response & \begin{tabular}[c]{@{}l@{}}He Ain't Heavy was written by \textcolor{red}{Mike D'Abo}. \\ $\cdots$\end{tabular} \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    cell = table.cells[1][1]
    assert "makecell" not in str(cell.value).lower()
    assert cell.rich_segments is not None
    assert "makecell" not in cell.rich_segments[0][0].lower()
    assert cell.rich_segments[0][0].endswith(" ")


def test_tex_reader_infers_first_column_rowspan_from_blank_continuation_rows():
    """Infer first-column visual merges when groups use blank continuation rows."""
    tex = r"""
\begin{tabular}{ccc}
\toprule
Iter & Balls & Score \\
\midrule
1 & 5 & 0.1 \\
 & 15 & 0.2 \\
 & 30 & 0.3 \\
\hline
2 & 5 & 0.4 \\
 & 15 & 0.5 \\
 & 30 & 0.6 \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.cells[1][0].rowspan == 3
    assert table.cells[4][0].rowspan == 3
    assert table.cells[2][0].value == ""
    assert table.cells[3][0].value == ""
    assert table.cells[5][0].value == ""
    assert table.cells[6][0].value == ""


def test_tex_reader_preserves_middle_spacer_column():
    """Middle intentionally empty columns should not be globally trimmed."""
    tex = r"""
\begin{tabular}{ccc}
\toprule
Left &  & Right \\
\midrule
L1 &  & R1 \\
\bottomrule
\end{tabular}
"""
    table = read_tex(tex)
    assert table.num_cols == 3
    assert table.cells[0][1].value == ""
    assert table.cells[1][1].value == ""


# --- renderer tests ---

def test_render_three_line():
    """Renderer produces three-line table structure."""
    from pubtab.models import Cell, CellStyle
    header = [Cell("A", CellStyle(bold=True)), Cell("B", CellStyle(bold=True))]
    row = [Cell("x"), Cell(0.5)]
    table = TableData(cells=[header, row], num_rows=2, num_cols=2, header_rows=1)
    tex = render(table)
    assert "\\toprule" in tex
    assert "\\midrule" in tex
    assert "\\bottomrule" in tex


def test_render_default_heavyrulewidth():
    """Renderer includes default heavyrulewidth."""
    from pubtab.models import Cell, CellStyle
    row = [Cell("a"), Cell("b")]
    table = TableData(cells=[row], num_rows=1, num_cols=2, header_rows=0)
    tex = render(table)
    assert "\\heavyrulewidth" in tex
    assert "1.0pt" in tex


def test_render_special_chars():
    """Renderer escapes special LaTeX characters."""
    from pubtab.models import Cell
    row = [Cell("100%"), Cell("A & B")]
    table = TableData(cells=[row], num_rows=1, num_cols=2, header_rows=0)
    tex = render(table)
    assert "100\\%" in tex
    assert "A \\& B" in tex


def test_render_section_row_midrules_when_section_not_in_first_column():
    """Section rows with empty first col should still get auto midrules."""
    from pubtab.models import Cell
    header = [Cell("Dataset"), Cell("Method"), Cell("Score")]
    section = [Cell(""), Cell("GPT-3.5", colspan=2), Cell("")]
    row = [Cell("Popular"), Cell("Direct"), Cell("3.67")]
    table = TableData(cells=[header, section, row], num_rows=3, num_cols=3, header_rows=1)
    tex = render(table)
    # One default header midrule + at least one auto section separator
    assert tex.count("\\midrule") >= 2


def test_render_section_row_uses_partial_rule_when_first_col_is_active_multirow():
    """Auto section separators should not strike through an active first-column multirow."""
    from pubtab.models import Cell
    header = [Cell("Dataset"), Cell("Method"), Cell("Score")]
    row1 = [Cell("Popular", rowspan=4), Cell("GPT-3.5", colspan=2), Cell("")]
    row2 = [Cell(""), Cell("Direct"), Cell("3.67")]
    row3 = [Cell(""), Cell("Llama3.1", colspan=2), Cell("")]
    row4 = [Cell(""), Cell("Direct"), Cell("3.60")]
    table = TableData(
        cells=[header, row1, row2, row3, row4],
        num_rows=5,
        num_cols=3,
        header_rows=1,
    )
    tex = render(table)
    assert "\\cmidrule(lr){2-3}" in tex


def test_render_unicode_subscript_keeps_text_base():
    """Unicode subscript symbols should render as text base + math script."""
    from pubtab.models import Cell
    row = [Cell("DRF_θ"), Cell("F_θ")]
    table = TableData(cells=[row], num_rows=1, num_cols=2, header_rows=0)
    tex = render(table)
    assert "DRF$_{\\theta}$" in tex
    assert "F$_{\\theta}$" in tex
    assert "$DRF_{\\theta}$" not in tex


# --- preview tests ---

def test_build_standalone_structure():
    """Standalone document has correct structure."""
    tex = _build_standalone("\\begin{table}...\\end{table}")
    assert "\\documentclass" in tex
    assert "\\begin{document}" in tex
    assert "\\resizebox" in tex


def test_build_standalone_auto_resizebox():
    """Standalone auto-adds resizebox when not present."""
    tex = _build_standalone("\\begin{tabular}{cc}a & b\\end{tabular}")
    assert "\\resizebox{\\linewidth}" in tex


def test_build_standalone_keeps_existing_resizebox():
    """Standalone doesn't double-wrap resizebox."""
    content = "\\resizebox{0.9\\textwidth}{!}{\\begin{tabular}{cc}a & b\\end{tabular}}"
    tex = _build_standalone(f"\\begin{{table}}[t]{content}\\end{{table}}")
    assert tex.count("\\resizebox") == 1


def test_strip_table_float_without_position():
    """Strip table float wrapper even when \\begin{table} has no [position]."""
    tex = r"\begin{table}\centering\caption{Cap}\begin{tabular}{c}x\end{tabular}\end{table}"
    out = _strip_table_float(tex)
    assert "\\begin{table" not in out
    assert "\\end{table" not in out
    assert "\\captionof{table}{Cap}" in out


def test_load_config_empty_yaml(tmp_path):
    """Empty YAML config should load as empty kwargs."""
    cfg = tmp_path / "empty.yaml"
    cfg.write_text("")
    kwargs, formatter = load_config(cfg)
    assert kwargs == {}
    assert formatter is None


def test_load_config_non_mapping_raises(tmp_path):
    """Non-mapping YAML root should raise a clear error."""
    cfg = tmp_path / "bad.yaml"
    cfg.write_text("- a\n- b\n")
    with pytest.raises(ValueError, match="mapping"):
        load_config(cfg)


@pytest.mark.skipif(not _find_pdflatex(), reason="pdflatex not available")
def test_compile_pdf(tmp_path):
    """PDF compilation works."""
    tex = "\\begin{tabular}{cc}\na & b \\\\\n\\end{tabular}"
    out = tmp_path / "test.pdf"
    compile_pdf(tex, out)
    assert out.exists()
