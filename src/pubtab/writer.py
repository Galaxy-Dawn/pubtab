"""Excel writer — converts TableData to .xlsx."""

from __future__ import annotations

from pathlib import Path
from typing import List, Union

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.rich_text import InlineFont

from .models import TableData
from .utils import _latex_color_to_hex


def _write_sheet(ws, table: TableData) -> None:
    """Write a single TableData to an openpyxl worksheet."""
    merged: set[tuple[int, int]] = set()

    for r, row in enumerate(table.cells):
        col_offset = 0
        for c, cell in enumerate(row):
            excel_row = r + 1
            excel_col = c + 1

            if (r, c) in merged:
                if cell.value == "" or cell.value is None:
                    continue
                # Cell has content but overlaps a previous merge — remove from merged
                merged.discard((r, c))

            # Determine value
            if cell.rich_segments:
                rt = CellRichText()
                for seg in cell.rich_segments:
                    seg_text, seg_color = seg[0], seg[1]
                    seg_bold = seg[2] if len(seg) > 2 else cell.style.bold
                    seg_italic = seg[3] if len(seg) > 3 else cell.style.italic
                    seg_underline = seg[4] if len(seg) > 4 else cell.style.underline
                    ifont = InlineFont(
                        b=seg_bold or None,
                        i=seg_italic or None,
                        u="single" if seg_underline else None,
                        # Explicitly write black for uncolored segments so openpyxl
                        # preserves segment boundaries when the cell also has a bg fill.
                        # Without this, the first colorless segment merges into the next
                        # colored segment on readback (e.g. "00.0" + green "±0.00" → all green).
                        color="FF" + seg_color.lstrip("#") if seg_color else "FF000000",
                    )
                    rt.append(TextBlock(ifont, seg_text))
                val = rt
            elif cell.style.diagbox:
                val = " / ".join(cell.style.diagbox)
            else:
                val = cell.value if cell.value != "" else None

            target = ws.cell(row=excel_row, column=excel_col)
            # Defensive guard for malformed overlap cases: writing to a
            # non-master merged cell raises read-only errors in openpyxl.
            if isinstance(target, MergedCell):
                if val in ("", None):
                    continue
                # Keep pipeline running even if source merge semantics are broken.
                continue
            target.value = val

            # Font styling (skip for rich text cells)
            if not cell.rich_segments:
                font_kwargs = dict(
                    bold=cell.style.bold,
                    italic=cell.style.italic,
                    underline="single" if cell.style.underline else None,
                )
                if cell.style.color:
                    hex_color = _latex_color_to_hex(cell.style.color)
                    if hex_color:
                        font_kwargs["color"] = hex_color.lstrip("#")
                ws.cell(row=excel_row, column=excel_col).font = Font(**font_kwargs)

            # Background color
            if cell.style.bg_color:
                hex_bg = _latex_color_to_hex(cell.style.bg_color)
                if hex_bg:
                    bg_val = hex_bg.lstrip("#")
                    ws.cell(row=excel_row, column=excel_col).fill = PatternFill(
                        start_color=bg_val, end_color=bg_val, fill_type="solid",
                    )

            # Alignment (wrap_text for multi-line cells, rotation for rotatebox)
            wrap = isinstance(val, str) and "\n" in val
            ws.cell(row=excel_row, column=excel_col).alignment = Alignment(
                horizontal=_align_map(cell.style.alignment),
                vertical="center",
                wrap_text=wrap,
                text_rotation=cell.style.rotation if cell.style.rotation else 0,
            )

            # Merge cells (cap rowspan to avoid overlapping with next content cell)
            if cell.rowspan > 1 or cell.colspan > 1:
                actual_rowspan = min(cell.rowspan, len(table.cells) - r)
                if actual_rowspan > 1:
                    for dr in range(1, actual_rowspan):
                        nr = r + dr
                        if nr < len(table.cells) and table.cells[nr][c].value not in ("", None):
                            actual_rowspan = dr
                            break
                end_row = excel_row + actual_rowspan - 1
                end_col = excel_col + cell.colspan - 1
                if end_row > excel_row or end_col > excel_col:
                    ws.merge_cells(
                        start_row=excel_row, start_column=excel_col,
                        end_row=end_row, end_column=end_col,
                    )
                for mr in range(r, r + actual_rowspan):
                    for mc in range(c, c + cell.colspan):
                        if (mr, mc) != (r, c):
                            merged.add((mr, mc))

    # Auto-adjust column widths based on content
    for col_idx in range(1, (ws.max_column or table.num_cols) + 1):
        max_len = 0
        for row_idx in range(1, (ws.max_row or len(table.cells)) + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                cell_len = max(len(str(line)) for line in str(val).split("\n"))
                max_len = max(max_len, cell_len)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 3, 4)


def write_excel(table: TableData, output: Union[str, Path]) -> Path:
    """Write TableData to an Excel file.

    Args:
        table: The table data to write.
        output: Output .xlsx file path.

    Returns:
        Path to the generated file.
    """
    output = Path(output)
    wb = Workbook()
    _write_sheet(wb.active, table)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    return output


def write_excel_multi(tables: List[TableData], output: Union[str, Path]) -> Path:
    """Write multiple TableData objects to separate sheets in one Excel file.

    Args:
        tables: List of TableData to write.
        output: Output .xlsx file path.

    Returns:
        Path to the generated file.
    """
    output = Path(output)
    wb = Workbook()
    for i, table in enumerate(tables):
        if i == 0:
            ws = wb.active
            ws.title = "Table 1"
        else:
            ws = wb.create_sheet(title=f"Table {i + 1}")
        _write_sheet(ws, table)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    return output


def _align_map(latex_align: str) -> str:
    """Map LaTeX alignment to Excel alignment."""
    a = latex_align[0] if latex_align else "c"
    return {"l": "left", "r": "right", "c": "center"}.get(a, "center")
