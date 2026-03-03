"""Excel file reader — converts .xlsx/.xls to TableData."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional, Union

from .models import Cell, CellStyle, TableData


def _extract_rich_segments(raw_value) -> Optional[tuple]:
    """Extract rich_segments from CellRichText, or return None."""
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
    except ImportError:
        return None
    if not isinstance(raw_value, CellRichText):
        return None
    segs = []
    has_formatting = False
    for block in raw_value:
        if isinstance(block, str):
            segs.append((block, None, False, False, False))
        elif isinstance(block, TextBlock):
            text = block.text or ""
            font = block.font
            color = None
            bold = bool(font.b) if font.b is not None else False
            italic = bool(font.i) if font.i is not None else False
            underline = bool(font.u) if font.u else False
            if font.color and font.color.rgb and isinstance(font.color.rgb, str):
                rgb = font.color.rgb
                if rgb.startswith("FF") and len(rgb) == 8:
                    color = f"#{rgb[2:]}"
            if color or bold or italic or underline:
                has_formatting = True
            segs.append((text, color, bold, italic, underline))
    if not has_formatting or len(segs) < 2:
        return None
    return tuple(segs)


def _excel_fmt_to_python(fmt: str) -> Optional[str]:
    """Convert Excel number format to Python format spec."""
    # Decimal: 0.00, #.000, etc.
    m = re.match(r'^[#0]*\.([0]+)$', fmt)
    if m:
        return f'.{len(m.group(1))}f'
    # Percentage: 0%, 0.00%, etc.
    m = re.match(r'^[#0]*\.?([0]*)%$', fmt)
    if m:
        decimals = len(m.group(1))
        return f'.{decimals}%' if decimals else '.0%'
    return None


def read_excel(
    path: Union[str, Path],
    sheet: Optional[Union[str, int]] = None,
    header_rows: Optional[int] = None,
) -> TableData:
    """Read an Excel file and return a TableData object.

    Args:
        path: Path to .xlsx or .xls file.
        sheet: Sheet name or 0-based index. Defaults to active sheet.
        header_rows: Number of rows to treat as header. Auto-detected if None.

    Returns:
        TableData with cell values, styles, and merge info.
    """
    path = Path(path)
    if path.suffix.lower() == ".xls":
        return _read_xls(path, sheet, header_rows)
    return _read_xlsx(path, sheet, header_rows)


def _read_xlsx(
    path: Path,
    sheet: Optional[Union[str, int]],
    header_rows: Optional[int],
) -> TableData:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, rich_text=True)
    if sheet is None:
        ws = wb.active
    elif isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        ws = wb[sheet]

    # Build merged cell map: (row, col) -> (master_row, master_col, rowspan, colspan)
    merge_map: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    for rng in ws.merged_cells.ranges:
        r1, c1, r2, c2 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        rs, cs = r2 - r1 + 1, c2 - c1 + 1
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                merge_map[(r, c)] = (r1, c1, rs, cs)

    num_rows = ws.max_row or 0
    num_cols = ws.max_column or 0
    cells: list[list[Cell]] = []

    for r in range(1, num_rows + 1):
        row_cells: list[Cell] = []
        for c in range(1, num_cols + 1):
            if (r, c) in merge_map:
                mr, mc, rs, cs = merge_map[(r, c)]
                if (r, c) != (mr, mc):
                    # Non-master merged cell: emit empty placeholder
                    row_cells.append(Cell(value="", style=CellStyle()))
                    continue
                rowspan, colspan = rs, cs
            else:
                rowspan, colspan = 1, 1

            cell = ws.cell(row=r, column=c)
            style = _extract_xlsx_style(cell)
            raw_value = cell.value
            rich_segments = None

            # Extract rich text segments from CellRichText
            rich_segments = _extract_rich_segments(raw_value)
            if rich_segments is not None:
                value = "".join(seg[0] for seg in rich_segments)
            else:
                value = raw_value if raw_value is not None else ""

            # Auto-detect diagbox: "X / Y" pattern — only top-left corner with text labels
            if isinstance(value, str) and " / " in value and r == 1 and c == 1:
                parts = value.split(" / ", 1)
                _numeric = all(p.strip().replace('.','',1).replace('-','',1).lstrip('-').isdigit() or p.strip() == '--' for p in parts)
                if not _numeric:
                    style = CellStyle(
                        bold=style.bold, italic=style.italic, underline=style.underline,
                        color=style.color, bg_color=style.bg_color, alignment=style.alignment,
                        fmt=style.fmt, diagbox=parts, rotation=style.rotation,
                    )
                    value = ""

            # Auto-detect raw LaTeX: if cell value contains \command patterns,
            # mark as raw_latex so the renderer passes it through unescaped
            if isinstance(value, str) and re.search(r'\\[a-zA-Z]', value):
                style = CellStyle(
                    raw_latex=True,
                    bold=style.bold, italic=style.italic, underline=style.underline,
                    color=style.color, bg_color=style.bg_color, alignment=style.alignment,
                    fmt=style.fmt, diagbox=style.diagbox, rotation=style.rotation,
                )

            row_cells.append(Cell(value=value, style=style, rowspan=rowspan, colspan=colspan, rich_segments=rich_segments))
        cells.append(row_cells)

    # Strip trailing empty rows
    while len(cells) > 1 and all(not c.value and c.value != 0 for c in cells[-1]):
        cells.pop()
    num_rows = len(cells)

    if header_rows is None:
        header_rows = max((c.rowspan for c in cells[0]), default=1) if cells else 1
        r = 1
        while r < header_rows and r < len(cells):
            for cell in cells[r]:
                header_rows = max(header_rows, r + cell.rowspan)
            r += 1

    return TableData(cells=cells, num_rows=num_rows, num_cols=num_cols, header_rows=header_rows)


def _extract_xlsx_style(cell) -> CellStyle:
    """Extract style from an openpyxl cell."""
    font = cell.font
    align = cell.alignment

    bold = bool(font.bold)
    italic = bool(font.italic)
    underline = bool(font.underline)
    color = None
    if font.color and font.color.rgb and isinstance(font.color.rgb, str) and font.color.rgb != "00000000":
        color = f"#{font.color.rgb[-6:]}"

    bg_color = None
    fill = cell.fill
    if fill and fill.fill_type and fill.fill_type != "none":
        if fill.fgColor and fill.fgColor.rgb and isinstance(fill.fgColor.rgb, str):
            rgb = fill.fgColor.rgb[-6:]
            bg_color = f"#{rgb}"

    alignment = "center"
    if align and align.horizontal:
        alignment = align.horizontal

    rotation = 0
    if align and align.text_rotation:
        rotation = align.text_rotation

    fmt = None
    if cell.number_format and cell.number_format != "General":
        fmt = _excel_fmt_to_python(cell.number_format)

    return CellStyle(bold=bold, italic=italic, underline=underline, color=color, bg_color=bg_color, alignment=alignment, fmt=fmt, rotation=rotation)


def _read_xls(
    path: Path,
    sheet: Optional[Union[str, int]],
    header_rows: Optional[int],
) -> TableData:
    import xlrd

    wb = xlrd.open_workbook(str(path), formatting_info=True)
    if sheet is None:
        ws = wb.sheet_by_index(0)
    elif isinstance(sheet, int):
        ws = wb.sheet_by_index(sheet)
    else:
        ws = wb.sheet_by_name(sheet)

    # Build merged cell map
    merge_map: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    for rlo, rhi, clo, chi in ws.merged_cells:
        rs, cs = rhi - rlo, chi - clo
        for r in range(rlo, rhi):
            for c in range(clo, chi):
                merge_map[(r, c)] = (rlo, clo, rs, cs)

    num_rows = ws.nrows
    num_cols = ws.ncols
    cells: list[list[Cell]] = []

    for r in range(num_rows):
        row_cells: list[Cell] = []
        for c in range(num_cols):
            if (r, c) in merge_map:
                mr, mc, rs, cs_val = merge_map[(r, c)]
                if (r, c) != (mr, mc):
                    row_cells.append(Cell(value="", style=CellStyle()))
                    continue
                rowspan, colspan = rs, cs_val
            else:
                rowspan, colspan = 1, 1

            value = ws.cell_value(r, c)
            if value == "":
                value = ""

            xf_idx = ws.cell_xf_index(r, c)
            style = _extract_xls_style(wb, xf_idx)
            row_cells.append(Cell(value=value, style=style, rowspan=rowspan, colspan=colspan))
        cells.append(row_cells)

    if header_rows is None:
        header_rows = max((c.rowspan for c in cells[0]), default=1) if cells else 1
        r = 1
        while r < header_rows and r < len(cells):
            for cell in cells[r]:
                header_rows = max(header_rows, r + cell.rowspan)
            r += 1

    return TableData(cells=cells, num_rows=num_rows, num_cols=num_cols, header_rows=header_rows)


def _extract_xls_style(wb, xf_idx: int) -> CellStyle:
    """Extract style from xlrd workbook using XF index."""
    try:
        xf = wb.xf_list[xf_idx]
        font = wb.font_list[xf.font_index]
        bold = bool(font.bold)
        italic = bool(font.italic)
    except (IndexError, AttributeError):
        bold, italic = False, False

    return CellStyle(bold=bold, italic=italic)
